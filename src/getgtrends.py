#!/usr/bin/env python3
# coding: utf-8
import csv
import collections
import json
from multiprocessing import Pool
import os
import pprint
import re
import sys
import time
import traceback
import feedparser
import pandas as pd
from pandas import Series, DataFrame

from pytrends.request import TrendReq
import mecaboperate as mec
import openpyxl as opx


def gtrend_getfeed(urls):
    """
    GoogleトレンドのRSSフィードを取得する。
    """
    try:
        results=[]
        keys = [
                 'title','ht_approx_traffic','published_parsed'
               ]
        feed = feedparser.parse(urls)
        for x in feed.entries:
            if x is not None:
               if hasattr(x, 'title') and \
                  hasattr(x, 'ht_approx_traffic') and \
                  hasattr(x, 'published_parsed'):
                  values = [
                             x.title,
                             x.ht_approx_traffic.replace('+','').replace(',',''),
                             time.strftime('%Y-%m-%d %H:%M:%S', x.published_parsed)
                           ]
                  o = dict(zip(keys,values))
                  results.append(o)
        if results:
           results = sorted(results, key=lambda x:x['ht_approx_traffic'],reverse=True)
           return results
        else:
            return
    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))

def gtrend_getvalue(kw_list,output_file,timeframe):
    """
    ライブラリを使用してGoogleTrendsからデータを取得する。
    #pytrends ref https://pypi.org/project/pytrends/#interest-by-region
    """
    try:
        sp = kw_list[0]
        pytrends = TrendReq(hl='ja-JP', tz=360)
        pytrends.build_payload(kw_list, cat=0, timeframe=timeframe, geo='JP', gprop='')
        #関連キーワード
        trendsdata = pytrends.related_queries()
        o = output_file
        s = sp + 'query'
        exportdata(trendsdata,o,s,1)
        #関連トピック
        trendsdata = pytrends.related_topics()
        s = sp + 'topic'
        exportdata(trendsdata,o,s,1)
        #地域別の関心
        trendsdata = pytrends.interest_by_region(resolution='REGION', inc_low_vol=True, inc_geo_code=False)
        s = sp + 'region'
        exportdata(trendsdata,o,s,0)
        #時系列
        trendsdata = pytrends.interest_over_time()
        s = sp + 'overtime'
        exportdata(trendsdata,o,s,0)
        #サジェスト　
        trendsdata = pytrends.suggestions(sp)
        s = sp + 'suggestions'
        suggest_to_excel(trendsdata,o,s)

        #注目キーワード
        #trendsword = pytrends.trending_searches(pn='united_states') #アメリカ
        #trendsword = pytrends.trending_searches(pn='japan') #日本
        #s = "trendword"
        #f = exportdata(trendsword,o,s,0)

    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))

def suggest_to_excel(trendsdata,output_file_name,sheetname):
    """
    サジェストをEXCELに書込み
    """
    #TODO EXCELにシートを追記してサジェスト内容を追記
    try:
        o = output_file_name + ".xlsx"
        wb = opx.load_workbook(o)
        ws = wb.create_sheet(sheetname)
        i = 0
        for sw in trendsdata:
            if sw :
               r = sw['title']
               if r is not None:
                  s = 'A' + str(i+1)
                  ws[s] = r
                  i+=1
        wb.save(o)
    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))

def fileexport(output_data,output_file_name,sheetname,mode):
    """
    モードを指定して出力ファイルを切替
    """
    try:
        if mode == 'excel':
            o = output_file_name +".xlsx"
            xlsf = os.path.isfile(o)
            if xlsf:
                with pd.ExcelWriter(o,engine="openpyxl", mode="a") as writer:
                    output_data.to_excel(writer, sheet_name=sheetname)
            else:
                with pd.ExcelWriter(o,engine="openpyxl") as writer:
                    output_data.to_excel(writer, sheet_name=sheetname)
        elif mode == 'csv':
            output_data.to_csv(output_file_name + "_" + sheetname + ".csv")
        else:
            print(output_data)
    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))

def exportdata(trendsdata,output_file_name,sheetname,data_type):
    """
    ファイル出力
    """
    try:
        fm = 'excel'
        if data_type == 1:
            data = [i for i in trendsdata.values()]
            for d1 in data:
                dictval = d1.values()
            for pandasdf in dictval:
                o = pandasdf
                if o is not None:
                   if not o.empty: 
                      fileexport(o,output_file_name,sheetname,fm)
        else:
            o = trendsdata
            fileexport(o,output_file_name,sheetname,fm)
    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))

def importfile(input_file):
    """
    ファイルから検索キーワードを取得

    """
    try:
        result=[[]]
        with open(input_file, "r") as f:
            data = csv.reader(f)
            result =[ r for r in data if r ]
        return result
    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))



def getfeedword(input_file):
    """
    RSSフィードを保存したファイルからタイトルと説明文を取得する。
    """
    try:
        o = []
        with open(input_file, newline = "") as f:
             df = json.load(f)
             for i in range(len(df)):
                 for j in range(len(df[i])):
                     if df[i] is not None or type(df[i]) is not 'NoneType':
                        t = df[i][j]['title']
                        s = df[i][j]['description']
                        o +=[t,s]
             return o
    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))

def getnumwords(targetword):
    """
    単語数を数えて上位からソートする
    """
    try:
        tw = collections.Counter(targetword)
        results = []
        keys = ['word','word_count']
        for word, cnt in sorted(tw.items(),key=lambda x: x[1], reverse=True):
            values = [word,cnt]
            w = dict(zip(keys,values))
            if w:
               results.append(w)
        if results:
           return results
        else:
           return
    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))


def main():
    """
    主処理
    
    """
    try:
        start_t = time.perf_counter()

        input_file = sys.argv[1]
        output_file = sys.argv[2]

        #trendsurls = 'https://trends.google.co.jp/trends/trendingsearches/daily/rss?geo=JP'
        #gw = gtrend_getfeed(trendsurls)
        #print(gw)

        #キーワードファイル取得
        kw_lists = importfile(input_file)
        for kw in kw_lists:
            if kw:
                #tframe='today 5-y'
                tframe='2018-01-01 2018-12-31'
                #5秒待機してからGoogleTorends結果取得
                time.sleep(5)
                gtrend_getvalue(kw,output_file,tframe)

        """feedword = getfeedword(input_file)
        wl = 3
        word = []
        p = re.compile('^[0-9]+$')
        for l in feedword:
            if len(l) > wl :
               wc = mec.word_tokenaize(l)
               o = [i for i in wc if not re.match(p,i)]
               word.extend(o)
        
        fw = getnumwords(word)
        #print(fw)"""


        end_t = time.perf_counter()
        process_time = end_t - start_t
        print('処理時間は:{0}秒です。'.format(process_time))
    except Exception as e:
        t, v, tb = sys.exc_info()
        print(traceback.format_exception(t,v,tb))
        print(traceback.format_tb(e.__traceback__))

if __name__ == '__main__':
   main()
